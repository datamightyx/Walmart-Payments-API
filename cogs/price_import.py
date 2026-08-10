# -*- coding: utf-8 -*-
"""Імпорт цін з xlsx (аркуш "COGS": Product, amazon ASIN, SKU, Item ID,
Total price/unit) у версійоване сховище цін.

Правила визначення effective_from:
  - SKU бачимо вперше (немає жодної ціни в БД)  -> SENTINEL_DATE (1900-01-01),
    щоб ціна діяла "завжди" і покривала вже завантажені історичні звіти.
  - SKU вже має ціну, і нова ціна відрізняється  -> effective_from = переданий
    параметр або сьогодні (нова версія, стара лишається для минулих дат).
  - Ціна не змінилась  -> нічого не пишемо.

dry_run=True повертає список змін без запису — для попереднього перегляду
перед підтвердженням у Streamlit.
"""

from __future__ import annotations

import sqlite3
from dataclasses import dataclass
from datetime import date
from pathlib import Path

import openpyxl

from . import store

SENTINEL_DATE = store.SENTINEL_DATE

REQUIRED_COLUMNS = {"Product", "SKU", "Total price/unit"}


class PriceImportError(RuntimeError):
    """xlsx не схожий на очікуваний прайс-файл."""


@dataclass
class PriceImportChange:
    sku: str
    product_name: str
    action: str  # "new" | "changed" | "unchanged"
    old_price: float | None
    new_price: float
    effective_from: str  # "" якщо action == "unchanged"


def _to_item_id(raw) -> str | None:
    if raw is None or raw == "":
        return None
    if isinstance(raw, (int, float)):
        return str(int(raw))
    return str(raw).strip()


def _read_rows(path: Path, sheet: str) -> list[dict]:
    wb = openpyxl.load_workbook(path, data_only=True)
    if sheet not in wb.sheetnames:
        raise PriceImportError(
            f"Аркуш {sheet!r} не знайдено у {path.name}. Доступні: {wb.sheetnames}"
        )
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    missing = REQUIRED_COLUMNS - set(header)
    if missing:
        raise PriceImportError(f"У аркуші {sheet!r} немає колонок: {sorted(missing)}")

    idx = {name: i for i, name in enumerate(header) if name}

    out: list[dict] = []
    for raw_row in rows[1:]:
        sku_cell = raw_row[idx["SKU"]] if idx["SKU"] < len(raw_row) else None
        price_cell = raw_row[idx["Total price/unit"]] if idx["Total price/unit"] < len(raw_row) else None
        if sku_cell is None or price_cell is None:
            continue
        product_cell = raw_row[idx["Product"]] if "Product" in idx and idx["Product"] < len(raw_row) else None
        item_id_cell = raw_row[idx["Item ID"]] if "Item ID" in idx and idx["Item ID"] < len(raw_row) else None
        asin_cell = raw_row[idx["amazon ASIN"]] if "amazon ASIN" in idx and idx["amazon ASIN"] < len(raw_row) else None
        out.append({
            "sku": str(sku_cell).strip(),
            "product_name": str(product_cell).strip() if product_cell else "",
            "item_id": _to_item_id(item_id_cell),
            "asin": str(asin_cell).strip() if asin_cell else None,
            "unit_cost": float(price_cell),
        })
    return out


def import_from_xlsx(
    conn: sqlite3.Connection,
    path: str | Path,
    sheet: str = "COGS",
    effective_from: date | None = None,
    source: str | None = None,
    dry_run: bool = False,
) -> list[PriceImportChange]:
    path = Path(path)
    rows = _read_rows(path, sheet)
    source = source or f"import:{path.name}"

    changes: list[PriceImportChange] = []
    for row in rows:
        sku = row["sku"]
        is_new_sku = not store.has_any_price(conn, sku)
        current = None if is_new_sku else store.latest_price(conn, sku, date.max)
        old_price = current["unit_cost"] if current else None

        if is_new_sku:
            eff = SENTINEL_DATE
            action = "new"
        elif old_price is not None and abs(old_price - row["unit_cost"]) < 1e-9:
            eff = None
            action = "unchanged"
        else:
            eff = effective_from or date.today()
            action = "changed"

        changes.append(PriceImportChange(
            sku=sku,
            product_name=row["product_name"],
            action=action,
            old_price=old_price,
            new_price=row["unit_cost"],
            effective_from=eff.isoformat() if eff else "",
        ))

        if not dry_run and action != "unchanged":
            store.insert_price_version(
                conn,
                sku=sku,
                item_id=row["item_id"],
                asin=row["asin"],
                product_name=row["product_name"],
                unit_cost=row["unit_cost"],
                effective_from=eff,
                source=source,
            )

    return changes
